"""
Generate test data files for regression testing.

Creates CSV, XLS, XLSX, JSON, and XML files for all 17 regression test configurations.
Always overwrites existing files to ensure fresh test data.
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
import xlwt


class TestFileGenerator:
    """Generates test data files for regression testing."""

    def __init__(self, base_dir: str = "/app/data/regression"):
        """Initialize generator with base directory."""
        self.base_dir = Path(base_dir)
        self.csv_dir = self.base_dir / "csv"
        self.xls_dir = self.base_dir / "xls"
        self.xlsx_dir = self.base_dir / "xlsx"
        self.json_dir = self.base_dir / "json"
        self.xml_dir = self.base_dir / "xml"

        # Create directories if they don't exist
        self.csv_dir.mkdir(parents=True, exist_ok=True)
        self.xls_dir.mkdir(parents=True, exist_ok=True)
        self.xlsx_dir.mkdir(parents=True, exist_ok=True)
        self.json_dir.mkdir(parents=True, exist_ok=True)
        self.xml_dir.mkdir(parents=True, exist_ok=True)

    def generate_csv_files(self):
        """Generate missing CSV test files."""
        print("Generating CSV files...")

        # Test 1: Strategy1_Products (5 records)
        products_file = self.csv_dir / "Strategy1_Products_20260101T120000.csv"
        with open(products_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['product_name', 'price', 'quantity', 'category'])
            writer.writeheader()
            writer.writerows([
                {'product_name': 'Laptop', 'price': '999.99', 'quantity': '10', 'category': 'Electronics'},
                {'product_name': 'Mouse', 'price': '19.99', 'quantity': '50', 'category': 'Electronics'},
                {'product_name': 'Keyboard', 'price': '79.99', 'quantity': '30', 'category': 'Electronics'},
                {'product_name': 'Monitor', 'price': '299.99', 'quantity': '15', 'category': 'Electronics'},
                {'product_name': 'Webcam', 'price': '89.99', 'quantity': '25', 'category': 'Electronics'},
            ])
        print(f"  Created: {products_file}")

        # Test 2: Strategy2_Products (3 records - same columns as Strategy1 + extras)
        # This tests Strategy 2: extra columns (supplier, sku) should be ignored
        products_strategy2_file = self.csv_dir / "Strategy2_Products_20260101T130000.csv"
        with open(products_strategy2_file, 'w', newline='', encoding='utf-8') as f:
            # Same columns as Strategy1 + extra columns
            writer = csv.DictWriter(f, fieldnames=['product_name', 'price', 'quantity', 'category', 'supplier', 'sku'])
            writer.writeheader()
            writer.writerows([
                {'product_name': 'Headphones', 'price': '149.99', 'quantity': '40', 'category': 'Electronics', 'supplier': 'AudioTech', 'sku': 'SKU-001'},
                {'product_name': 'USB Cable', 'price': '9.99', 'quantity': '100', 'category': 'Electronics', 'supplier': 'CableCo', 'sku': 'SKU-002'},
                {'product_name': 'Desk Lamp', 'price': '39.99', 'quantity': '20', 'category': 'Electronics', 'supplier': 'LightWorks', 'sku': 'SKU-003'},
            ])
        print(f"  Created: {products_strategy2_file}")

        # Test 3: Strategy3_Orders (4 records - for strict validation)
        orders_strategy3_file = self.csv_dir / "Strategy3_Orders_20260101T140000.csv"
        with open(orders_strategy3_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['order_id', 'customer', 'total'])
            writer.writeheader()
            writer.writerows([
                {'order_id': 'ORD-001', 'customer': 'Alice Johnson', 'total': '150.00'},
                {'order_id': 'ORD-002', 'customer': 'Bob Smith', 'total': '275.50'},
                {'order_id': 'ORD-003', 'customer': 'Carol Davis', 'total': '99.99'},
                {'order_id': 'ORD-004', 'customer': 'David Wilson', 'total': '450.00'},
            ])
        print(f"  Created: {orders_strategy3_file}")

        # Test 4: MetadataFilename (2 records - label extracted from filename)
        metadata_filename_file = self.csv_dir / "MetadataFilename_20260101T150000.csv"
        with open(metadata_filename_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['item', 'value'])
            writer.writeheader()
            writer.writerows([
                {'item': 'Widget', 'value': '100'},
                {'item': 'Gadget', 'value': '200'},
            ])
        print(f"  Created: {metadata_filename_file}")

        # Test 5: EmptyFile (headers only)
        empty_file = self.csv_dir / "EmptyFile_20260101T160000.csv"
        with open(empty_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['field1', 'field2', 'field3'])
            writer.writeheader()
        print(f"  Created: {empty_file}")

        # Test 6: MalformedData (2 records with some odd data)
        malformed_file = self.csv_dir / "MalformedData_20260101T170000.csv"
        with open(malformed_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['field_a', 'field_b'])
            writer.writeheader()
            writer.writerows([
                {'field_a': 'Value1', 'field_b': '123'},
                {'field_a': 'Value2', 'field_b': '456'},
            ])
        print(f"  Created: {malformed_file}")

    def generate_xls_files(self):
        """Generate XLS test files using xlwt."""
        print("Generating XLS files...")

        # Test 7: Strategy1_Inventory (7 records)
        inventory_file = self.xls_dir / "Strategy1_Inventory_20260101T110000.xls"
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Inventory')

        # Headers
        headers = ['item_code', 'item_name', 'stock_qty', 'warehouse', 'category']
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        # Data
        inventory_data = [
            ['ITM001', 'Widget A', 150, 'Warehouse 1', 'Widgets'],
            ['ITM002', 'Widget B', 200, 'Warehouse 1', 'Widgets'],
            ['ITM003', 'Gadget X', 75, 'Warehouse 2', 'Gadgets'],
            ['ITM004', 'Gadget Y', 100, 'Warehouse 2', 'Gadgets'],
            ['ITM005', 'Tool Z', 50, 'Warehouse 3', 'Tools'],
            ['ITM006', 'Part A', 300, 'Warehouse 1', 'Parts'],
            ['ITM007', 'Part B', 250, 'Warehouse 3', 'Parts'],
        ]
        for row, data in enumerate(inventory_data, start=1):
            for col, value in enumerate(data):
                ws.write(row, col, value)

        wb.save(str(inventory_file))
        print(f"  Created: {inventory_file}")

        # Test 8: MetadataContent (4 records with category for metadata)
        metadata_file = self.xls_dir / "MetadataContent_20260101T120000.xls"
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Data')

        headers = ['category', 'product', 'quantity', 'price']
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        metadata_data = [
            ['Electronics', 'Laptop', 5, 999.99],
            ['Electronics', 'Mouse', 20, 19.99],
            ['Electronics', 'Keyboard', 15, 79.99],
            ['Electronics', 'Monitor', 8, 299.99],
        ]
        for row, data in enumerate(metadata_data, start=1):
            for col, value in enumerate(data):
                ws.write(row, col, value)

        wb.save(str(metadata_file))
        print(f"  Created: {metadata_file}")

        # Test 9: MultipleSheets (3 records - note: only first sheet is read)
        multisheet_file = self.xls_dir / "MultipleSheets_20260101T130000.xls"
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet1')

        headers = ['sheet_name', 'data_value', 'timestamp']
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        sheet_data = [
            ['Sheet1', 'Value1', '2026-01-01 10:00:00'],
            ['Sheet1', 'Value2', '2026-01-01 11:00:00'],
            ['Sheet1', 'Value3', '2026-01-01 12:00:00'],
        ]
        for row, data in enumerate(sheet_data, start=1):
            for col, value in enumerate(data):
                ws.write(row, col, value)

        # Add a second sheet (won't be imported, but demonstrates multi-sheet file)
        ws2 = wb.add_sheet('Sheet2')
        ws2.write(0, 0, 'This sheet is ignored')

        wb.save(str(multisheet_file))
        print(f"  Created: {multisheet_file}")

    def generate_xlsx_files(self):
        """Generate XLSX test files using openpyxl."""
        print("Generating XLSX files...")

        # Test 10: Strategy2_Sales (10 records)
        sales_file = self.xlsx_dir / "Strategy2_Sales_20260101T140000.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sales'

        headers = ['sale_id', 'product', 'amount', 'date', 'region']
        ws.append(headers)

        sales_data = [
            [1001, 'Product A', 500.00, '2026-01-01', 'North'],
            [1002, 'Product B', 750.00, '2026-01-01', 'South'],
            [1003, 'Product C', 1200.00, '2026-01-01', 'East'],
            [1004, 'Product A', 550.00, '2026-01-01', 'West'],
            [1005, 'Product D', 300.00, '2026-01-01', 'North'],
            [1006, 'Product B', 800.00, '2026-01-01', 'South'],
            [1007, 'Product C', 1100.00, '2026-01-01', 'East'],
            [1008, 'Product E', 425.00, '2026-01-01', 'West'],
            [1009, 'Product A', 575.00, '2026-01-01', 'North'],
            [1010, 'Product D', 350.00, '2026-01-01', 'South'],
        ]
        for row in sales_data:
            ws.append(row)

        wb.save(str(sales_file))
        print(f"  Created: {sales_file}")

        # Test 11: DateContent (5 records with report_date for date extraction)
        datecontent_file = self.xlsx_dir / "DateContent_20260101T150000.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'

        headers = ['report_date', 'metric_name', 'value', 'notes']
        ws.append(headers)

        date_data = [
            ['2026-01-01', 'Revenue', 50000, 'Q1 start'],
            ['2026-01-01', 'Expenses', 35000, 'Within budget'],
            ['2026-01-01', 'Profit', 15000, 'Target met'],
            ['2026-01-01', 'Customers', 1250, 'Growing'],
            ['2026-01-01', 'Orders', 875, 'Steady'],
        ]
        for row in date_data:
            ws.append(row)

        wb.save(str(datecontent_file))
        print(f"  Created: {datecontent_file}")

        # Test 12: LargeFile (1000 records)
        large_file = self.xlsx_dir / "LargeFile_20260101T160000.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'LargeData'

        headers = ['row_id', 'data_col1', 'data_col2', 'data_col3']
        ws.append(headers)

        print(f"  Generating 1000 records for large file...")
        for i in range(1, 1001):
            ws.append([
                i,
                f'Data_{i}_Col1',
                f'Data_{i}_Col2',
                f'Data_{i}_Col3'
            ])

        wb.save(str(large_file))
        print(f"  Created: {large_file}")

    def generate_json_files(self):
        """Generate JSON test files."""
        print("Generating JSON files...")

        # Test 13: JSON Array format (6 records in array)
        array_file = self.json_dir / "ArrayFormat_20260104T120000.json"
        array_data = [
            {"product_id": 1, "name": "Product A", "price": 19.99},
            {"product_id": 2, "name": "Product B", "price": 29.99},
            {"product_id": 3, "name": "Product C", "price": 39.99},
            {"product_id": 4, "name": "Product D", "price": 49.99},
            {"product_id": 5, "name": "Product E", "price": 59.99},
            {"product_id": 6, "name": "Product F", "price": 69.99},
        ]
        with open(array_file, 'w', encoding='utf-8') as f:
            json.dump(array_data, f, indent=2)
        print(f"  Created: {array_file}")

        # Test 14: JSON Object format (1 record)
        object_file = self.json_dir / "ObjectFormat_20260104T130000.json"
        object_data = {
            "transaction_id": "TXN-12345",
            "customer": "John Doe",
            "amount": 150.75,
            "currency": "USD"
        }
        with open(object_file, 'w', encoding='utf-8') as f:
            json.dump(object_data, f, indent=2)
        print(f"  Created: {object_file}")

        # Test 15: JSON Nested Objects (1 record with nested structure)
        nested_file = self.json_dir / "NestedObjects_20260104T140000.json"
        nested_data = {
            "order_id": "ORD-999",
            "customer": {
                "name": "Jane Smith",
                "email": "jane@example.com"
            },
            "items": ["Item1", "Item2", "Item3"]
        }
        with open(nested_file, 'w', encoding='utf-8') as f:
            json.dump(nested_data, f, indent=2)
        print(f"  Created: {nested_file}")

    def generate_xml_files(self):
        """Generate XML test files."""
        print("Generating XML files...")

        # Test 16: XML Structured format (3 records)
        structured_file = self.xml_dir / "StructuredXML_20260105T120000.xml"
        structured_xml = '''<?xml version="1.0"?>
<products>
  <product>
    <id>101</id>
    <name>Widget Pro</name>
    <category>Tools</category>
  </product>
  <product>
    <id>102</id>
    <name>Gadget Max</name>
    <category>Electronics</category>
  </product>
  <product>
    <id>103</id>
    <name>Tool Supreme</name>
    <category>Hardware</category>
  </product>
</products>'''
        with open(structured_file, 'w', encoding='utf-8') as f:
            f.write(structured_xml)
        print(f"  Created: {structured_file}")

        # Test 17: XML Blob format (2 records - complex nested structure)
        blob_file = self.xml_dir / "BlobXML_20260105T130000.xml"
        blob_xml = '''<?xml version="1.0"?>
<complex_document>
  <header>
    <title>Quarterly Report</title>
    <date>2026-01-05</date>
  </header>
  <body>
    <section id="1">
      <paragraph>Lorem ipsum dolor sit amet, consectetur adipiscing elit.</paragraph>
      <chart type="bar">
        <data points="10,20,30,40,50"/>
      </chart>
    </section>
  </body>
</complex_document>'''
        with open(blob_file, 'w', encoding='utf-8') as f:
            f.write(blob_xml)
        print(f"  Created: {blob_file}")

    def generate_all(self):
        """Generate all test files."""
        print("\n" + "=" * 60)
        print("REGRESSION TEST FILE GENERATOR")
        print("=" * 60)
        print(f"Base directory: {self.base_dir}")
        print()

        self.generate_csv_files()
        print()
        self.generate_xls_files()
        print()
        self.generate_xlsx_files()
        print()
        self.generate_json_files()
        print()
        self.generate_xml_files()

        print()
        print("=" * 60)
        print("FILE GENERATION COMPLETE")
        print("=" * 60)
        print()
        print("Summary:")
        print(f"  CSV files:  6 created in {self.csv_dir}")
        print(f"  XLS files:  3 created in {self.xls_dir}")
        print(f"  XLSX files: 3 created in {self.xlsx_dir}")
        print(f"  JSON files: 3 created in {self.json_dir}")
        print(f"  XML files:  2 created in {self.xml_dir}")
        print(f"  Total:      17 test files created")
        print()
        print("All test files have been generated/overwritten successfully!")
        print()


def main():
    """Main entry point."""
    generator = TestFileGenerator()
    generator.generate_all()


if __name__ == "__main__":
    main()
