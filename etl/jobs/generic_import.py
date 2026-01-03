"""
Generic file import job driven by dba.timportconfig configuration.

This job reads import configurations from the database and dynamically imports
files (CSV, XLS, XLSX, JSON, XML) into target tables with support for:
- Dynamic column management (ALTER TABLE to add new columns)
- Metadata extraction from filename, file content, or static values
- Date extraction with configurable formats
- File archiving after processing

Usage:
    # Run from command line
    python etl/jobs/generic_import.py --config-id 1

    # Or programmatically
    from etl.jobs.generic_import import GenericImportJob
    job = GenericImportJob(config_id=1)
    job.run()
"""

import argparse
import csv
import json
import os
import re
import shutil
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree

from common.db_utils import db_connection, db_transaction, fetch_dict, bulk_insert
from common.logging_utils import get_logger
from etl.base.etl_job import BaseETLJob
from etl.loaders.postgres_loader import PostgresLoader


@dataclass
class ImportConfig:
    """Import configuration from dba.timportconfig."""
    config_id: int
    config_name: str
    datasource: str
    datasettype: str
    source_directory: str
    archive_directory: str
    file_pattern: str
    file_type: str
    metadata_label_source: str
    metadata_label_location: Optional[str]
    dateconfig: str
    datelocation: Optional[str]
    dateformat: Optional[str]
    delimiter: Optional[str]
    target_table: str
    importstrategyid: int
    is_active: bool


@dataclass
class ImportStrategy:
    """Import strategy from dba.timportstrategy."""
    importstrategyid: int
    name: str
    description: Optional[str]


class ConfigNotFoundError(Exception):
    """Raised when import configuration is not found."""
    pass


class ImportValidationError(Exception):
    """Raised when import validation fails."""
    pass


class FileExtractor(ABC):
    """Abstract base class for file extractors."""

    @abstractmethod
    def extract(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract data from file as list of dictionaries."""
        pass


class CSVExtractor(FileExtractor):
    """Extract data from CSV files."""

    def __init__(self, delimiter: str = ','):
        self.delimiter = delimiter if delimiter else ','

    def extract(self, file_path: Path) -> List[Dict[str, Any]]:
        records = []
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=self.delimiter)
            for row in reader:
                records.append(dict(row))
        return records


class ExcelExtractor(FileExtractor):
    """Extract data from XLS/XLSX files."""

    def __init__(self, file_type: str):
        self.file_type = file_type

    def extract(self, file_path: Path) -> List[Dict[str, Any]]:
        if self.file_type == 'XLSX':
            return self._extract_xlsx(file_path)
        else:
            return self._extract_xls(file_path)

    def _extract_xlsx(self, file_path: Path) -> List[Dict[str, Any]]:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else f'column_{i}' for i, h in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            record = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = value
            records.append(record)

        wb.close()
        return records

    def _extract_xls(self, file_path: Path) -> List[Dict[str, Any]]:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        if ws.nrows == 0:
            return []

        headers = [str(ws.cell_value(0, i)).strip() or f'column_{i}' for i in range(ws.ncols)]
        records = []
        for row_idx in range(1, ws.nrows):
            record = {}
            for col_idx in range(ws.ncols):
                value = ws.cell_value(row_idx, col_idx)
                record[headers[col_idx]] = value
            records.append(record)

        return records


class JSONExtractor(FileExtractor):
    """Extract data from JSON files."""

    def extract(self, file_path: Path) -> List[Dict[str, Any]]:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            return [{'raw_data': json.dumps(data), 'source_file': file_path.name}]
        else:
            return [{'raw_data': json.dumps(data), 'source_file': file_path.name}]


class XMLExtractor(FileExtractor):
    """Extract data from XML files."""

    def extract(self, file_path: Path) -> List[Dict[str, Any]]:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        try:
            root = ElementTree.fromstring(content)
            records = self._parse_element(root)
            if records:
                return records
        except ElementTree.ParseError:
            pass

        return [{'raw_data': content, 'source_file': file_path.name}]

    def _parse_element(self, element) -> List[Dict[str, Any]]:
        """Try to parse XML element into records."""
        records = []
        for child in element:
            record = {}
            for subchild in child:
                record[subchild.tag] = subchild.text
            if record:
                records.append(record)
        return records


class SchemaManager:
    """Manages dynamic table schema changes."""

    def __init__(self, logger=None):
        self.logger = logger or get_logger(self.__class__.__name__)

    def get_table_columns(self, schema: str, table: str) -> List[str]:
        """Get existing column names for a table."""
        query = """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = %s
            ORDER BY ordinal_position
        """
        results = fetch_dict(query, (schema, table))
        return [row['column_name'] for row in results]

    def table_exists(self, schema: str, table: str) -> bool:
        """Check if table exists."""
        query = """
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_schema = %s AND table_name = %s
            ) as exists
        """
        results = fetch_dict(query, (schema, table))
        return results[0]['exists'] if results else False

    def add_column(self, schema: str, table: str, column_name: str, column_type: str = 'TEXT'):
        """Add a new column to the table."""
        safe_column = self._sanitize_identifier(column_name)
        with db_transaction() as cursor:
            cursor.execute(f'ALTER TABLE {schema}.{table} ADD COLUMN IF NOT EXISTS "{safe_column}" {column_type}')
        self.logger.info(f"Added column '{safe_column}' to {schema}.{table}")

    def _sanitize_identifier(self, name: str) -> str:
        """Sanitize column name for SQL."""
        sanitized = re.sub(r'[^\w]', '_', name.lower())
        sanitized = re.sub(r'_+', '_', sanitized)
        sanitized = sanitized.strip('_')
        if sanitized[0].isdigit():
            sanitized = 'col_' + sanitized
        return sanitized

    def _infer_column_type(self, values: List[Any]) -> str:
        """
        Infer PostgreSQL column type from sample values.

        Analyzes multiple values to determine the most appropriate type.
        Returns TEXT as fallback for mixed or unknown types.
        """
        # Filter out None values
        non_null_values = [v for v in values if v is not None and v != '']

        if not non_null_values:
            return 'TEXT'

        # Track type observations
        has_int = False
        has_float = False
        has_bool = False
        max_length = 0

        for value in non_null_values:
            # Convert to string to check actual value
            str_value = str(value).strip()
            max_length = max(max_length, len(str_value))

            # Check for boolean
            if isinstance(value, bool) or str_value.lower() in ('true', 'false', 't', 'f'):
                has_bool = True
                continue

            # Check for numeric types
            try:
                if isinstance(value, (int, float)):
                    num_value = value
                else:
                    num_value = float(str_value)

                # Check if it's actually an integer
                if isinstance(value, int) or (isinstance(num_value, float) and num_value.is_integer()):
                    has_int = True
                else:
                    has_float = True
            except (ValueError, TypeError):
                # Not a number, will default to text
                pass

        # Determine best type based on observations
        if has_bool and not has_int and not has_float:
            return 'BOOLEAN'
        elif has_float:
            return 'NUMERIC(12,2)'
        elif has_int:
            # Check max value to determine if we need BIGINT
            max_val = max([abs(int(v)) for v in non_null_values if str(v).lstrip('-').isdigit()], default=0)
            if max_val > 2147483647:  # Max INT value
                return 'BIGINT'
            return 'INTEGER'
        else:
            # Text type - determine appropriate length
            if max_length <= 50:
                return 'VARCHAR(50)'
            elif max_length <= 100:
                return 'VARCHAR(100)'
            elif max_length <= 255:
                return 'VARCHAR(255)'
            else:
                return 'TEXT'

    def create_table_from_records(
        self,
        schema: str,
        table: str,
        sample_records: List[Dict[str, Any]],
        max_samples: int = 5
    ):
        """
        Create a new table dynamically based on sample records.

        Analyzes up to max_samples records to infer column names and types.
        Creates table following feeds schema conventions:
        - {table}id SERIAL PRIMARY KEY
        - datasetid INT FK to dba.tdataset
        - Business columns (inferred from data)
        - created_date TIMESTAMP
        - created_by VARCHAR(50)

        Args:
            schema: Schema name (e.g., 'feeds')
            table: Table name (e.g., 'products')
            sample_records: List of sample records to analyze
            max_samples: Number of records to analyze (default 5)
        """
        if not sample_records:
            raise ImportValidationError("Cannot create table from empty record set")

        # Analyze first N records
        samples_to_analyze = sample_records[:max_samples]

        # Collect all column names across samples
        all_columns = set()
        for record in samples_to_analyze:
            all_columns.update(record.keys())

        # Remove system/metadata columns that shouldn't be in table
        excluded_columns = {
            'datasetid', 'created_date', 'created_by', 'modified_date', 'modified_by',
            'source_file', 'metadata_label', 'file_date'
        }
        business_columns = sorted(all_columns - excluded_columns)

        # Gather sample values for each column
        column_samples = {col: [] for col in business_columns}
        for record in samples_to_analyze:
            for col in business_columns:
                if col in record:
                    column_samples[col].append(record[col])

        # Infer types for each column
        column_definitions = []
        for col in business_columns:
            safe_col = self._sanitize_identifier(col)
            col_type = self._infer_column_type(column_samples[col])
            column_definitions.append(f'    {safe_col} {col_type}')

        # Build CREATE TABLE statement
        pk_column = f'{table}id'
        column_lines = ',\n'.join(column_definitions)
        create_sql = f"""
CREATE TABLE {schema}.{table} (
    {pk_column} SERIAL PRIMARY KEY,
    datasetid INT REFERENCES dba.tdataset(datasetid),
{column_lines},
    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    created_by VARCHAR(50)
);
"""

        # Execute table creation
        with db_transaction() as cursor:
            cursor.execute(create_sql)

        self.logger.info(f"Created table {schema}.{table} with {len(business_columns)} business columns")

        # Grant permissions
        grant_sql = f"""
GRANT SELECT, INSERT, UPDATE, DELETE ON {schema}.{table} TO app_rw;
GRANT SELECT ON {schema}.{table} TO app_ro;
GRANT USAGE, SELECT ON SEQUENCE {schema}.{pk_column}_seq TO app_rw;
"""
        with db_transaction() as cursor:
            cursor.execute(grant_sql)

        self.logger.info(f"Granted permissions on {schema}.{table}")


class GenericImportJob(BaseETLJob):
    """
    Generic import job driven by database configuration.

    Reads configuration from dba.timportconfig and imports files
    according to the specified settings.
    """

    def __init__(
        self,
        config_id: int,
        run_date: Optional[date] = None,
        dry_run: bool = False,
        **kwargs
    ):
        """
        Initialize generic import job.

        Args:
            config_id: ID from dba.timportconfig.config_id
            run_date: Date for this run (defaults to today)
            dry_run: If True, don't actually load data
            **kwargs: Additional arguments passed to BaseETLJob

        Raises:
            ConfigNotFoundError: If config_id is not found in database
        """
        self.config_id = config_id
        self.import_config: Optional[ImportConfig] = None
        self.strategy: Optional[ImportStrategy] = None
        self.matched_files: List[Path] = []
        self.schema_manager = SchemaManager()
        self._temp_logger = get_logger('GenericImportJob')

        import_config = self._load_config(config_id)
        if import_config is None:
            raise ConfigNotFoundError(
                f"Config ID {config_id} not found. Check dba.timportconfig table."
            )

        self.import_config = import_config
        self.strategy = self._load_strategy(import_config.importstrategyid)

        super().__init__(
            run_date=run_date or date.today(),
            dataset_type=import_config.datasettype,
            data_source=import_config.datasource,
            dry_run=dry_run,
            **kwargs
        )

        self.loader = PostgresLoader(schema=self._get_target_schema())

    def _load_config(self, config_id: int) -> Optional[ImportConfig]:
        """Load import configuration from database."""
        query = """
            SELECT
                config_id, config_name, datasource, datasettype,
                source_directory, archive_directory, file_pattern, file_type,
                metadata_label_source, metadata_label_location,
                dateconfig, datelocation, dateformat, delimiter,
                target_table, importstrategyid, is_active
            FROM dba.timportconfig
            WHERE config_id = %s AND is_active = TRUE
        """
        results = fetch_dict(query, (config_id,))
        if not results:
            return None

        row = results[0]
        return ImportConfig(
            config_id=row['config_id'],
            config_name=row['config_name'],
            datasource=row['datasource'],
            datasettype=row['datasettype'],
            source_directory=row['source_directory'],
            archive_directory=row['archive_directory'],
            file_pattern=row['file_pattern'],
            file_type=row['file_type'],
            metadata_label_source=row['metadata_label_source'],
            metadata_label_location=row['metadata_label_location'],
            dateconfig=row['dateconfig'],
            datelocation=row['datelocation'],
            dateformat=row['dateformat'],
            delimiter=row['delimiter'],
            target_table=row['target_table'],
            importstrategyid=row['importstrategyid'],
            is_active=row['is_active']
        )

    def _load_strategy(self, strategy_id: int) -> ImportStrategy:
        """Load import strategy from database."""
        query = """
            SELECT importstrategyid, name, description
            FROM dba.timportstrategy
            WHERE importstrategyid = %s
        """
        results = fetch_dict(query, (strategy_id,))
        if not results:
            return ImportStrategy(
                importstrategyid=strategy_id,
                name='Unknown',
                description=None
            )

        row = results[0]
        return ImportStrategy(
            importstrategyid=row['importstrategyid'],
            name=row['name'],
            description=row.get('description')
        )

    def _get_target_schema(self) -> str:
        """Extract schema from target_table (format: schema.table)."""
        if '.' in self.import_config.target_table:
            return self.import_config.target_table.split('.')[0]
        return 'feeds'

    def _get_target_table(self) -> str:
        """Extract table name from target_table (format: schema.table)."""
        if '.' in self.import_config.target_table:
            return self.import_config.target_table.split('.')[1]
        return self.import_config.target_table

    def _get_extractor(self) -> FileExtractor:
        """Get appropriate file extractor based on file type."""
        file_type = self.import_config.file_type.upper()
        if file_type == 'CSV':
            # Note: The config.delimiter is for filename parsing, not CSV parsing
            # CSV files always use comma as the delimiter
            return CSVExtractor(delimiter=',')
        elif file_type in ('XLS', 'XLSX'):
            return ExcelExtractor(file_type=file_type)
        elif file_type == 'JSON':
            return JSONExtractor()
        elif file_type == 'XML':
            return XMLExtractor()
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def _find_matching_files(self) -> List[Path]:
        """Find files matching the configured pattern."""
        source_dir = Path(self.import_config.source_directory)
        if not source_dir.exists():
            if self.logger:
                self.logger.warning(f"Source directory does not exist: {source_dir}")
            return []

        pattern = self.import_config.file_pattern
        if self.logger:
            self.logger.debug(f"Scanning directory: {source_dir}")
            self.logger.debug(f"Using pattern: {pattern}")
        matched = []

        for file_path in source_dir.iterdir():
            if self.logger:
                self.logger.debug(f"Checking file: {file_path.name} (is_file: {file_path.is_file()})")
            if file_path.is_file():
                match_result = re.match(pattern, file_path.name)
                if self.logger:
                    self.logger.debug(f"  Pattern match result: {match_result}")
                if match_result:
                    matched.append(file_path)
                    if self.logger:
                        self.logger.debug(f"  MATCHED: {file_path.name}")

        if self.logger:
            self.logger.debug(f"Total matched files: {len(matched)}")
        return sorted(matched)

    def _extract_metadata_label(self, file_path: Path, records: List[Dict]) -> str:
        """Extract metadata label based on configuration."""
        source = self.import_config.metadata_label_source
        location = self.import_config.metadata_label_location

        if source == 'static':
            return location or 'Unknown'
        elif source == 'filename':
            if self.import_config.delimiter and location:
                parts = file_path.stem.split(self.import_config.delimiter)
                try:
                    index = int(location)
                    if 0 <= index < len(parts):
                        return parts[index]
                except (ValueError, IndexError):
                    pass
            return file_path.stem
        elif source == 'file_content':
            if records and location:
                first_record = records[0]
                if location in first_record:
                    return str(first_record[location])
            return file_path.stem

        return file_path.stem

    def _extract_date(self, file_path: Path, records: List[Dict]) -> date:
        """Extract date based on configuration."""
        source = self.import_config.dateconfig
        location = self.import_config.datelocation
        date_format = self.import_config.dateformat

        date_str = None

        if source == 'static':
            date_str = location
        elif source == 'filename':
            if self.import_config.delimiter and location:
                parts = file_path.stem.split(self.import_config.delimiter)
                try:
                    index = int(location)
                    if 0 <= index < len(parts):
                        date_str = parts[index]
                except (ValueError, IndexError):
                    pass
        elif source == 'file_content':
            if records and location:
                first_record = records[0]
                if location in first_record:
                    date_str = str(first_record[location])

        if date_str and date_format:
            try:
                py_format = self._convert_date_format(date_format)
                parsed = datetime.strptime(date_str, py_format)
                return parsed.date()
            except ValueError:
                pass

        return date.today()

    def _convert_date_format(self, format_str: str) -> str:
        """Convert Java-style date format to Python strptime format."""
        conversions = {
            'yyyy': '%Y',
            'yy': '%y',
            'MM': '%m',
            'dd': '%d',
            'HH': '%H',
            'mm': '%M',
            'ss': '%S',
            'T': 'T',
        }
        result = format_str
        for java_fmt, py_fmt in conversions.items():
            result = result.replace(java_fmt, py_fmt)
        return result

    def _normalize_column_names(self, records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Normalize column names to be SQL-safe."""
        if not records:
            return records

        normalized = []
        for record in records:
            new_record = {}
            for key, value in record.items():
                new_key = self.schema_manager._sanitize_identifier(key)
                new_record[new_key] = value
            normalized.append(new_record)

        return normalized

    def _apply_import_strategy(
        self,
        records: List[Dict[str, Any]],
        existing_columns: List[str]
    ) -> List[Dict[str, Any]]:
        """Apply import strategy to handle column mismatches."""
        if not records:
            return records

        source_columns = set(records[0].keys())
        existing_set = set(existing_columns)

        # System columns that should never be dynamically added to raw feeds tables
        # These are either managed by the loader or are internal metadata
        system_columns = {
            'datasetid', 'created_date', 'created_by', 'modified_date', 'modified_by',
            # Internal metadata columns added during extraction (should not go to feeds tables)
            'source_file', 'metadata_label', 'file_date'
        }
        comparable_existing = existing_set - system_columns

        # Only consider actual data columns from CSV, not system/metadata columns
        data_columns = source_columns - system_columns
        new_columns = data_columns - comparable_existing
        missing_columns = comparable_existing - data_columns

        strategy_id = self.strategy.importstrategyid

        if strategy_id == 1:
            # Only add actual new data columns from the CSV, not metadata columns
            schema = self._get_target_schema()
            table = self._get_target_table()
            for col in new_columns:
                self.schema_manager.add_column(schema, table, col)
            # Filter out metadata columns before returning
            allowed_system = {'datasetid', 'created_date', 'created_by', 'modified_date', 'modified_by'}
            valid_columns = data_columns | allowed_system
            filtered = []
            for record in records:
                filtered_record = {k: v for k, v in record.items() if k in valid_columns}
                filtered.append(filtered_record)
            return filtered

        elif strategy_id == 2:
            # Filter to only existing columns + allowed system columns (not metadata)
            allowed_system = {'datasetid', 'created_date', 'created_by', 'modified_date', 'modified_by'}
            valid_columns = (data_columns & comparable_existing) | allowed_system
            filtered = []
            for record in records:
                filtered_record = {k: v for k, v in record.items() if k in valid_columns}
                filtered.append(filtered_record)
            if new_columns:
                self.logger.info(f"Ignoring columns not in target table: {new_columns}")
            return filtered

        elif strategy_id == 3:
            if new_columns:
                raise ImportValidationError(
                    f"Source file contains columns not in target table: {new_columns}. "
                    f"Import strategy 3 requires all source columns to exist in target table."
                )
            # Filter out metadata columns
            allowed_system = {'datasetid', 'created_date', 'created_by', 'modified_date', 'modified_by'}
            valid_columns = data_columns | allowed_system
            filtered = []
            for record in records:
                filtered_record = {k: v for k, v in record.items() if k in valid_columns}
                filtered.append(filtered_record)
            return filtered

        return records

    def _ensure_reference_data(self):
        """
        Ensure datasource and datasettype exist in reference tables.

        Creates them automatically if they don't exist to support modular configuration.
        """
        try:
            # Ensure datasource exists
            self.logger.debug(f"Checking if datasource '{self.import_config.datasource}' exists...")
            check_source_query = "SELECT COUNT(*) as count FROM dba.tdatasource WHERE sourcename = %s"
            source_result = fetch_dict(check_source_query, (self.import_config.datasource,))
            self.logger.debug(f"Datasource query result: {source_result}")

            if source_result[0]['count'] == 0:
                self.logger.info(f"Datasource '{self.import_config.datasource}' does not exist - creating...")
                insert_source_query = """
                    INSERT INTO dba.tdatasource (sourcename, description, createdby)
                    VALUES (%s, %s, %s)
                """
                with db_transaction() as cursor:
                    cursor.execute(
                        insert_source_query,
                        (
                            self.import_config.datasource,
                            f'Auto-created for import config: {self.import_config.config_name}',
                            self.username
                        )
                    )
                self.logger.info(f"Created datasource: {self.import_config.datasource}")
            else:
                self.logger.debug(f"Datasource '{self.import_config.datasource}' already exists")

            # Ensure datasettype exists
            self.logger.debug(f"Checking if datasettype '{self.import_config.datasettype}' exists...")
            check_type_query = "SELECT COUNT(*) as count FROM dba.tdatasettype WHERE typename = %s"
            type_result = fetch_dict(check_type_query, (self.import_config.datasettype,))
            self.logger.debug(f"Datasettype query result: {type_result}")

            if type_result[0]['count'] == 0:
                self.logger.info(f"Datasettype '{self.import_config.datasettype}' does not exist - creating...")
                insert_type_query = """
                    INSERT INTO dba.tdatasettype (typename, description, createdby)
                    VALUES (%s, %s, %s)
                """
                with db_transaction() as cursor:
                    cursor.execute(
                        insert_type_query,
                        (
                            self.import_config.datasettype,
                            f'Auto-created for import config: {self.import_config.config_name}',
                            self.username
                        )
                    )
                self.logger.info(f"Created datasettype: {self.import_config.datasettype}")
            else:
                self.logger.debug(f"Datasettype '{self.import_config.datasettype}' already exists")

        except Exception as e:
            self.logger.error(f"Error ensuring reference data: {e}", exc_info=True)
            raise

    def setup(self):
        """Set up resources and validate configuration."""
        self.logger.info(
            f"Starting generic import with config: {self.import_config.config_name}",
            extra={
                'stepcounter': 'setup',
                'metadata': {
                    'config_id': self.config_id,
                    'strategy': self.strategy.name,
                    'file_type': self.import_config.file_type,
                    'target_table': self.import_config.target_table
                }
            }
        )

        self.matched_files = self._find_matching_files()

        if not self.matched_files:
            self.logger.warning(
                f"No files found matching pattern '{self.import_config.file_pattern}' "
                f"in '{self.import_config.source_directory}'"
            )

        self.logger.info(f"Found {len(self.matched_files)} file(s) to process")

    def extract(self) -> List[Dict[str, Any]]:
        """Extract data from all matched files."""
        if not self.matched_files:
            return []

        extractor = self._get_extractor()
        all_records = []

        for file_path in self.matched_files:
            self.logger.info(f"Extracting from: {file_path.name}")
            try:
                records = extractor.extract(file_path)
                metadata_label = self._extract_metadata_label(file_path, records)
                file_date = self._extract_date(file_path, records)

                for record in records:
                    record['_source_file'] = file_path.name
                    record['_metadata_label'] = metadata_label
                    record['_file_date'] = file_date

                all_records.extend(records)
                self.logger.info(f"Extracted {len(records)} records from {file_path.name}")

            except Exception as e:
                self.logger.error(f"Error extracting from {file_path.name}: {e}")
                raise

        return all_records

    def transform(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Transform extracted data."""
        if not data:
            return []

        transformed = self._normalize_column_names(data)

        for record in transformed:
            record['created_date'] = datetime.now()
            record['created_by'] = self.username

        schema = self._get_target_schema()
        table = self._get_target_table()

        if self.schema_manager.table_exists(schema, table):
            existing_columns = self.schema_manager.get_table_columns(schema, table)
            transformed = self._apply_import_strategy(transformed, existing_columns)
        else:
            # Table doesn't exist - create it dynamically from sample records
            self.logger.info(f"Target table {schema}.{table} does not exist - creating dynamically")
            self.schema_manager.create_table_from_records(schema, table, transformed, max_samples=5)
            self.logger.info(f"Successfully created table {schema}.{table}")

            # Now get the columns and apply import strategy
            existing_columns = self.schema_manager.get_table_columns(schema, table)
            transformed = self._apply_import_strategy(transformed, existing_columns)

        return transformed

    def load(self, data: List[Dict[str, Any]]):
        """Load transformed data to database."""
        if not data:
            self.logger.warning("No data to load")
            self.records_loaded = 0
            return

        # Remove any remaining metadata columns (normalized names, no underscore prefix)
        metadata_columns = {'source_file', 'metadata_label', 'file_date'}
        for record in data:
            for col in metadata_columns:
                record.pop(col, None)

        table = self._get_target_table()

        self.logger.info(f"Loading {len(data)} records to {self._get_target_schema()}.{table}")

        try:
            self.records_loaded = self.loader.load(
                table=table,
                data=data,
                dataset_id=self.dataset_id
            )
            self.logger.info(f"Successfully loaded {self.records_loaded} records")
        except Exception as e:
            self.logger.error(f"Load failed: {e}")
            raise

    def cleanup(self):
        """Archive processed files and update last_modified_at."""
        archive_dir = Path(self.import_config.archive_directory)
        if not archive_dir.exists():
            archive_dir.mkdir(parents=True, exist_ok=True)

        for file_path in self.matched_files:
            try:
                dest = archive_dir / file_path.name
                counter = 1
                while dest.exists():
                    stem = file_path.stem
                    suffix = file_path.suffix
                    dest = archive_dir / f"{stem}_{counter}{suffix}"
                    counter += 1

                shutil.move(str(file_path), str(dest))
                self.logger.info(f"Archived {file_path.name} to {dest}")
            except Exception as e:
                self.logger.error(f"Failed to archive {file_path.name}: {e}")

        try:
            with db_transaction() as cursor:
                cursor.execute(
                    "UPDATE dba.timportconfig SET last_modified_at = %s WHERE config_id = %s",
                    (datetime.now(), self.config_id)
                )
        except Exception as e:
            self.logger.warning(f"Failed to update last_modified_at: {e}")

        self.logger.info("Cleanup complete")

    def _extract_label_early(self) -> Optional[str]:
        """
        Extract dataset label early (before full extraction) based on import config.

        This is needed because dataset record is created before extract() is called,
        but we need the label from the import configuration.

        Returns:
            Extracted label or None if no files found
        """
        # Find matching files
        files = self._find_matching_files()
        if not files:
            return None

        # Use first file for label extraction
        file_path = files[0]
        source = self.import_config.metadata_label_source
        location = self.import_config.metadata_label_location

        # Handle static label
        if source == 'static':
            return location or 'Unknown'

        # Handle filename-based label
        elif source == 'filename':
            if self.import_config.delimiter and location:
                parts = file_path.stem.split(self.import_config.delimiter)
                try:
                    index = int(location)
                    if 0 <= index < len(parts):
                        return parts[index]
                except (ValueError, IndexError):
                    pass
            return file_path.stem

        # Handle file_content-based label (requires reading first record)
        elif source == 'file_content':
            if location:
                try:
                    extractor = self._get_extractor()
                    records = extractor.extract(file_path)
                    if records and location in records[0]:
                        return str(records[0][location])
                except Exception:
                    pass  # Fall back to filename
            return file_path.stem

        return file_path.stem

    def run(self):
        """
        Override run() to ensure reference data exists before parent's run().

        The parent run() calls _create_dataset_record() which requires
        datasource and datasettype to exist in reference tables.
        """
        # Use temp logger since self.logger isn't initialized yet
        temp_logger = get_logger(f'{self.__class__.__name__}.{self.run_uuid[:8]}')

        try:
            # Ensure datasource and datasettype exist before creating dataset record
            temp_logger.info("Ensuring reference data exists...")

            # Manually check and create since we don't have self.logger yet
            check_source_query = "SELECT COUNT(*) as count FROM dba.tdatasource WHERE sourcename = %s"
            source_result = fetch_dict(check_source_query, (self.import_config.datasource,))

            if source_result[0]['count'] == 0:
                temp_logger.info(f"Creating datasource: {self.import_config.datasource}")
                insert_source_query = "INSERT INTO dba.tdatasource (sourcename, description, createdby) VALUES (%s, %s, %s)"
                with db_transaction() as cursor:
                    cursor.execute(
                        insert_source_query,
                        (
                            self.import_config.datasource,
                            f'Auto-created for import config: {self.import_config.config_name}',
                            self.username
                        )
                    )

            check_type_query = "SELECT COUNT(*) as count FROM dba.tdatasettype WHERE typename = %s"
            type_result = fetch_dict(check_type_query, (self.import_config.datasettype,))

            if type_result[0]['count'] == 0:
                temp_logger.info(f"Creating datasettype: {self.import_config.datasettype}")
                insert_type_query = "INSERT INTO dba.tdatasettype (typename, description, createdby) VALUES (%s, %s, %s)"
                with db_transaction() as cursor:
                    cursor.execute(
                        insert_type_query,
                        (
                            self.import_config.datasettype,
                            f'Auto-created for import config: {self.import_config.config_name}',
                            self.username
                        )
                    )

            temp_logger.info("Reference data check complete")

            # Extract dataset label early based on import configuration
            temp_logger.info("Extracting dataset label from import configuration...")
            dataset_label = self._extract_label_early()
            if dataset_label:
                self.dataset_label = dataset_label
                temp_logger.info(f"Using dataset label: {dataset_label}")
            else:
                temp_logger.warning("No files found to extract label from - using default label pattern")

        except Exception as e:
            temp_logger.error(f"Error ensuring reference data: {e}", exc_info=True)
            raise

        # Now call parent run() which will create dataset record
        super().run()


def main():
    """Main entry point for command-line execution."""
    parser = argparse.ArgumentParser(
        description='Generic file import driven by database configuration'
    )
    parser.add_argument(
        '--config-id',
        type=int,
        required=True,
        help='Config ID from dba.timportconfig'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Run extraction and transformation without loading to database'
    )
    parser.add_argument(
        '--date',
        type=str,
        help='Run date in YYYY-MM-DD format (defaults to today)'
    )

    args = parser.parse_args()

    if not os.getenv('DB_URL'):
        print("ERROR: DB_URL environment variable not set")
        exit(1)

    run_date = None
    if args.date:
        try:
            run_date = datetime.strptime(args.date, '%Y-%m-%d').date()
        except ValueError:
            print(f"ERROR: Invalid date format: {args.date}. Use YYYY-MM-DD")
            exit(1)

    try:
        job = GenericImportJob(
            config_id=args.config_id,
            run_date=run_date,
            dry_run=args.dry_run
        )
    except ConfigNotFoundError as e:
        print(f"ERROR: {e}")
        exit(1)

    try:
        success = job.run()
        if success:
            print(f"\nGeneric import completed successfully")
            print(f"  - Config: {job.import_config.config_name}")
            print(f"  - Files processed: {len(job.matched_files)}")
            print(f"  - Extracted: {job.records_extracted} records")
            print(f"  - Transformed: {job.records_transformed} records")
            print(f"  - Loaded: {job.records_loaded} records")
            print(f"  - Run UUID: {job.run_uuid}")
            print(f"  - Dataset ID: {job.dataset_id}")
        else:
            print("\nGeneric import failed")
            exit(1)
    except Exception as e:
        print(f"\nGeneric import failed with error: {e}")
        exit(1)


if __name__ == "__main__":
    main()
